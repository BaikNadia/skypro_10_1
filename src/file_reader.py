import csv
import os

from openpyxl import load_workbook


def read_csv_file(file_path: str) -> list:
    """
    Считывает финансовые операции из CSV-файла.

    :param file_path: Путь к CSV-файлу.
    :return: Список словарей с транзакциями.
    """
    if not os.path.exists(file_path):
        return []  # Возвращаем пустой список, если файл не существует

    transactions = []
    try:
        with open(file_path, mode='r', encoding='utf-8') as file:
            reader = csv.DictReader(file, delimiter=';')  # Используем разделитель ';'
            for row in reader:
                transactions.append(row)
    except Exception:
        return []  # Возвращаем пустой список при ошибке чтения

    return transactions

def read_excel_file(file_path: str) -> list:
    """
    Считывает финансовые операции из Excel-файла.

    :param file_path: Путь к Excel-файлу.
    :return: Список словарей с транзакциями.
    """
    if not os.path.exists(file_path):
        print(f"Файл не найден: {file_path}")
        return []

    transactions = []
    try:
        workbook = load_workbook(filename=file_path)
        sheet = workbook.active

        # Получаем заголовки из первой строки
        headers = [cell.value for cell in sheet[1]]

        # Считываем данные из остальных строк
        for row in sheet.iter_rows(min_row=2, values_only=True):
            transaction = dict(zip(headers, row))
            transactions.append(transaction)
    except Exception as e:
        print(f"Ошибка чтения Excel-файла: {str(e)}")
        return []

    return transactions


def read_json_file():
    return None